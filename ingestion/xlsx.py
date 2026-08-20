"""Structure-preserving XLSX ingestion for spreadsheet knowledge sources."""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell, MergedCell  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

RowKind = Literal[
    "blank",
    "title",
    "metadata",
    "header",
    "data",
    "section",
    "note",
    "key_value",
]


@dataclass(frozen=True, slots=True)
class SpreadsheetCell:
    """One cell with its cached display value and optional formula."""

    column: int
    column_letter: str
    value: str | None
    formula: str | None


@dataclass(frozen=True, slots=True)
class SpreadsheetRow:
    """One physical worksheet row, including intentionally blank rows."""

    row_number: int
    kind: RowKind
    table_number: int | None
    cells: tuple[SpreadsheetCell, ...]
    headers: tuple[str | None, ...]
    hidden: bool


@dataclass(frozen=True, slots=True)
class ExtractedSheet:
    """Worksheet rows and workbook-level provenance."""

    source_path: str
    workbook_name: str
    sheet_name: str
    sheet_number: int
    state: str
    max_column: int
    rows: tuple[SpreadsheetRow, ...]


@dataclass(frozen=True, slots=True)
class ExtractedWorkbook:
    """Complete workbook extraction with sheet order preserved."""

    source_path: str
    workbook_name: str
    sheets: tuple[ExtractedSheet, ...]


@dataclass(frozen=True, slots=True)
class SpreadsheetChunk:
    """One indexable spreadsheet row with table and location context."""

    chunk_id: str
    source_path: str
    workbook_name: str
    sheet_name: str
    sheet_number: int
    row_number: int
    row_kind: RowKind
    table_number: int | None
    headers: tuple[str | None, ...]
    text: str


def _stringify(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _is_merged_row(worksheet: Worksheet, row_number: int) -> bool:
    return any(
        cell_range.min_row == row_number == cell_range.max_row
        for cell_range in worksheet.merged_cells.ranges
    )


def _looks_like_table_header(cells: tuple[Cell | MergedCell, ...]) -> bool:
    populated = [(index, cell) for index, cell in enumerate(cells) if cell.value is not None]
    if len(populated) < 2:
        return False
    indexes = [index for index, _ in populated]
    if indexes != list(range(indexes[0], indexes[-1] + 1)):
        return False
    return all(isinstance(cell, Cell) and cell.font.bold for _, cell in populated)


def _classify_row(
    worksheet: Worksheet,
    row_number: int,
    cells: tuple[Cell | MergedCell, ...],
    active_headers: tuple[str | None, ...],
) -> RowKind:
    populated = [cell for cell in cells if cell.value is not None]
    if not populated:
        return "blank"
    if _looks_like_table_header(cells):
        return "header"
    if _is_merged_row(worksheet, row_number):
        first_cell = populated[0]
        if row_number == 1:
            return "title"
        text = str(first_cell.value)
        if "Effective:" in text or "Version" in text:
            return "metadata"
        if isinstance(first_cell, Cell) and first_cell.font.bold:
            return "section"
        return "note"
    if active_headers:
        return "data"
    return "key_value"


def extract_xlsx(path: str | Path) -> ExtractedWorkbook:
    """Extract all physical rows while preserving formulas and cached values."""

    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(f"XLSX file does not exist: {workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Expected a .xlsx file, received: {workbook_path.name}")

    formula_workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    value_workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    source_path = workbook_path.as_posix()
    extracted_sheets: list[ExtractedSheet] = []

    for sheet_number, formula_sheet in enumerate(formula_workbook.worksheets, start=1):
        value_sheet = value_workbook[formula_sheet.title]
        rows: list[SpreadsheetRow] = []
        active_headers: tuple[str | None, ...] = ()
        table_number = 0

        for row_number in range(1, formula_sheet.max_row + 1):
            formula_cells = tuple(
                formula_sheet.cell(row=row_number, column=column)
                for column in range(1, formula_sheet.max_column + 1)
            )
            kind = _classify_row(formula_sheet, row_number, formula_cells, active_headers)
            if kind == "blank":
                active_headers = ()
            elif kind == "header":
                table_number += 1
                active_headers = tuple(_stringify(cell.value) for cell in formula_cells)

            cells: list[SpreadsheetCell] = []
            for column, formula_cell in enumerate(formula_cells, start=1):
                cached_value = value_sheet.cell(row=row_number, column=column).value
                effective_value = (
                    cached_value if formula_cell.data_type == "f" else formula_cell.value
                )
                formula = (
                    str(formula_cell.value)
                    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
                    else None
                )
                cells.append(
                    SpreadsheetCell(
                        column=column,
                        column_letter=get_column_letter(column),
                        value=_stringify(effective_value),
                        formula=formula,
                    )
                )

            rows.append(
                SpreadsheetRow(
                    row_number=row_number,
                    kind=kind,
                    table_number=table_number if kind in {"header", "data"} else None,
                    cells=tuple(cells),
                    headers=active_headers if kind in {"header", "data"} else (),
                    hidden=bool(formula_sheet.row_dimensions[row_number].hidden),
                )
            )

        extracted_sheets.append(
            ExtractedSheet(
                source_path=source_path,
                workbook_name=workbook_path.name,
                sheet_name=formula_sheet.title,
                sheet_number=sheet_number,
                state=formula_sheet.sheet_state,
                max_column=formula_sheet.max_column,
                rows=tuple(rows),
            )
        )

    formula_workbook.close()
    value_workbook.close()
    return ExtractedWorkbook(
        source_path=source_path,
        workbook_name=workbook_path.name,
        sheets=tuple(extracted_sheets),
    )


def _render_row(sheet: ExtractedSheet, row: SpreadsheetRow) -> str:
    context = [
        f"Workbook: {sheet.workbook_name}",
        f"Sheet: {sheet.sheet_name}",
        f"Row: {row.row_number}",
        f"Kind: {row.kind}",
    ]
    if row.table_number is not None:
        context.append(f"Table: {row.table_number}")

    if row.kind == "header":
        header_text = " | ".join(value or "[blank header]" for value in row.headers)
        context.append(f"Headers: {header_text}")
    elif row.kind == "data":
        values: list[str] = []
        for cell, header in zip(row.cells, row.headers, strict=True):
            label = header or f"Column {cell.column_letter}"
            value = cell.value if cell.value is not None else "[blank]"
            if cell.formula is not None:
                value = f"{value} [formula: {cell.formula}]"
            values.append(f"{label}={value}")
        context.append("Values: " + "; ".join(values))
    else:
        values = []
        for cell in row.cells:
            if cell.value is None and cell.formula is None:
                continue
            value = cell.value or "[no cached value]"
            if cell.formula is not None:
                value = f"{value} [formula: {cell.formula}]"
            values.append(f"{cell.column_letter}={value}")
        context.append("Values: " + "; ".join(values))

    return "\n".join(context)


def chunk_workbook(workbook: ExtractedWorkbook) -> list[SpreadsheetChunk]:
    """Create one searchable chunk per nonblank row with deterministic IDs."""

    chunks: list[SpreadsheetChunk] = []
    for sheet in workbook.sheets:
        for row in sheet.rows:
            if row.kind == "blank":
                continue
            text = _render_row(sheet, row)
            identity = json.dumps(
                {
                    "source_path": sheet.source_path,
                    "sheet_number": sheet.sheet_number,
                    "sheet_name": sheet.sheet_name,
                    "row_number": row.row_number,
                    "row_kind": row.kind,
                    "table_number": row.table_number,
                    "text": text,
                },
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            )
            chunks.append(
                SpreadsheetChunk(
                    chunk_id=hashlib.sha256(identity.encode("utf-8")).hexdigest()[:24],
                    source_path=sheet.source_path,
                    workbook_name=sheet.workbook_name,
                    sheet_name=sheet.sheet_name,
                    sheet_number=sheet.sheet_number,
                    row_number=row.row_number,
                    row_kind=row.kind,
                    table_number=row.table_number,
                    headers=row.headers,
                    text=text,
                )
            )
    return chunks


def ingest_xlsx(path: str | Path) -> list[SpreadsheetChunk]:
    """Run XLSX extraction and context-preserving chunking end to end."""

    return chunk_workbook(extract_xlsx(path))
